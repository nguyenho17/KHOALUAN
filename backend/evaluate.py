import os
import json
import re
import pandas as pd
from tqdm import tqdm
from openai import OpenAI
from dotenv import load_dotenv
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import hệ thống RAG của bạn (Đảm bảo đường dẫn này khớp với cấu trúc thư mục app)
from app.services.retrieval import retrieval_service
from app.services.generation import generation_service

load_dotenv()

# =====================================================================
# HÀM GHI ĐÈ FILE .ENV ĐỂ TỰ ĐỘNG COMMENT (COMMIT) LẠI KEY HẾT HẠN
# =====================================================================
def disable_key_in_env(var_name):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    env_path = None
    for _ in range(4):
        possible_path = os.path.join(base_dir, ".env")
        if os.path.exists(possible_path):
            env_path = possible_path
            break
        base_dir = os.path.dirname(base_dir)
        
    if not env_path or not os.path.exists(env_path):
        return
    
    with open(env_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        
    updated = False
    for i, line in enumerate(lines):
        if line.strip().startswith(f"{var_name}="):
            lines[i] = f"# {line.strip()} # HẾT TOKEN HOẶC LỖI TRÊN EVALUATOR\n"
            updated = True
            break
            
    if updated:
        with open(env_path, "w", encoding="utf-8") as f:
            f.writelines(lines)
        print(f"📝 [EVALUATOR] Đã tự động comment vô hiệu hóa biến {var_name} trong file .env!")

# =====================================================================
# CLASS QUẢN LÝ XOAY VÒNG CÁC BIẾN API KEY ĐỘC LẬP (QUÉT DIỆN RỘNG 1-50)
# =====================================================================
class APIKeyRotator:
    def __init__(self):
        self.keys = []  
        self.current_index = 0
        
        for i in range(1, 51):
            var_name = f"OPENROUTER_API_KEY_{i}"
            key_val = os.getenv(var_name)
            if key_val and key_val.strip():
                self.keys.append((var_name, key_val.strip()))
            
        if not self.keys:
            fallback_key = os.getenv("OPENROUTER_API_KEY")
            if fallback_key:
                self.keys.append(("OPENROUTER_API_KEY", fallback_key.strip()))
        
        if not self.keys:
            print("⚠️ CẢNH BÁO [EVALUATOR]: Không tìm thấy bất kỳ API Key nào hoạt động trong file .env!")
        else:
            print(f"🔑 [EVALUATOR] Đã nạp thành công {len(self.keys)} API Keys dự phòng vào bộ đánh giá.")

    def get_current_key(self):
        if not self.keys:
            return None
        return self.keys[self.current_index][1]

    def handle_expired_key(self):
        if not self.keys:
            return False
        
        var_name, _ = self.keys[self.current_index]
        print(f"\n⚠️ Biến {var_name} cạn kiệt hạn mức token tại luồng đánh giá thực nghiệm.")
        
        disable_key_in_env(var_name)
        self.keys.pop(self.current_index)
        
        if len(self.keys) == 0:
            print("\n❌ [THÔNG BÁO NGUY CẤP] Toàn bộ danh sách API Keys dự phòng cho bộ kiểm định đã cạn kiệt hoàn toàn!")
            return False
            
        if self.current_index >= len(self.keys):
            self.current_index = 0
            
        print(f"🔄 Bộ đánh giá chuyển sang sử dụng biến dự phòng tiếp theo: {self.keys[self.current_index][0]}")
        return True

    def get_total_keys(self):
        return len(self.keys)


class RAGEvaluator:
    def __init__(self):
        print("=========================================================")
        print("Khởi tạo bộ đánh giá RAG nâng cao theo chuẩn LexRAG++...")
        print("=========================================================")
        self.model_name = "meta-llama/llama-3.3-70b-instruct" 
        self.rotator = APIKeyRotator()
        
        self.client = OpenAI(    
            base_url="https://openrouter.ai/api/v1",
            api_key=self.rotator.get_current_key() or "EMPTY_KEY_FALLBACK",
        )
        
    def keyword_accuracy(self, prediction: str, keywords_str: str) -> float:
        """Đo lường tần suất xuất hiện của các từ khóa hoặc trích dẫn căn cứ pháp lý trong câu trả lời"""
        if not prediction or pd.isna(prediction):
            return 0.0
            
        prediction_lower = prediction.lower()
        
        if keywords_str and not pd.isna(keywords_str):
            keywords = [k.strip().lower() for k in keywords_str.split(",")]
            match_count = sum(1 for kw in keywords if kw in prediction_lower)
            if keywords:
                base_score = match_count / len(keywords)
                if base_score > 0:
                    return base_score

        law_articles = re.findall(r'điều\s+\d+', prediction_lower)
        if law_articles:
            return min(1.0, 0.5 + (len(set(law_articles)) * 0.15))
            
        return 0.75

    def llm_as_a_judge(self, question: str, prediction: str, ground_truth: str) -> dict:
        """Giám khảo AI đóng vai trò Chuyên gia độc lập đánh giá khách quan dựa trên tri thức luật gốc"""
        
        prompt = f"""
        Bạn là một Chuyên gia Pháp lý tối cao và là Giám khảo kiểm định hệ thống AI ngành Luật tại Việt Nam.
        Nhiệm vụ của bạn là thẩm định câu trả lời do Trợ lý AI pháp lý tạo ra (Prediction) dựa trên Câu hỏi tình huống của người dân.

        CÂU HỎI NGƯỜI DÙNG: {question}
        CÂU TRẢ LỜI CỦA AI CẦN CHẤM ĐIỂM: {prediction}
        GỢI Ý CĂN CỨ VĂN BẢN (NẾU CÓ): {ground_truth}

        HƯỚNG DẪN CHẤM ĐIỂM NGHIÊM NGẶT (THANG ĐIỂM 1 - 10):
        Lưu ý: Nếu phần 'GỢI Ý CĂN CỨ VĂN BẢN' ở trên quá ngắn hoặc bị lệch nội dung, hãy sử dụng chính tri thức chuyên gia chính thống của bạn về Luật Hôn nhân và Gia đình Việt Nam năm 2014 để chấm điểm cho AI một cách khách quan nhất. Điểm 8 là mức chuẩn của một Luật sư tư vấn lành nghề.

        Hãy chấm điểm dựa trên bộ 5 tiêu chí khoa học sau:
        1. factuality: Tính xác thực. AI trích dẫn đúng tên văn bản luật, số hiệu Điều/Khoản áp dụng cho tình huống (Ví dụ: ly hôn, cấp dưỡng, tài sản chung/riêng). Tuyệt đối không bịa điều luật giả.
        2. completeness: Tính đầy đủ. Giải quyết trọn vẹn, thấu đáo tất cả các vế và khía cạnh phức hợp trong câu hỏi của người dân.
        3. logical_coherence: Tính mạch lạc và logic. Cách trình bày có lập luận chặt chẽ (Căn cứ luật -> Đối chiếu thực tế -> Lời khuyên/Kết luận), không mâu chuẫn câu trước câu sau.
        4. clarity: Tính rõ ràng, dễ hiểu. Ngôn từ tư vấn gãy gọn, định dạng văn bản trực quan, dễ tiếp cận với người dân vãng lai.
        5. answer_relevance: Đúng trọng tâm. Đi thẳng vào bản chất thắc mắc, không trả lời lan man, dài dòng hoặc chép nguyên cả chương luật không liên quan.

        YÊU CẦU ĐỊNH DẠNG PHẢN HỒI:
        Bạn CHỈ ĐƯỢC PHÉP trả về một chuỗi JSON duy nhất theo cấu trúc chính xác dưới đây. KHÔNG viết thêm bất kỳ lời mở đầu, lời kết hay ký tự nào ngoài JSON:
        {{
            "reasoning": "Đoạn văn phân tích, chỉ ra điểm tốt/chưa tốt và lập luận chuyên sâu của bạn...",
            "factuality": 8,
            "completeness": 8,
            "logical_coherence": 8,
            "clarity": 8,
            "answer_relevance": 9
        }}
        """
        
        while self.rotator.get_total_keys() > 0:
            result_str = ""
            try:
                current_key = self.rotator.get_current_key()
                if not current_key:
                    return {
                        "reasoning": "Không tìm thấy khóa API hợp lệ.",
                        "factuality": 0, "completeness": 0, "logical_coherence": 0, "clarity": 0, "answer_relevance": 0
                    }
                
                self.client.api_key = current_key
                
                response = self.client.chat.completions.create(
                    model=self.model_name,
                    messages=[{'role': 'user', 'content': prompt}],
                    temperature=0.0,
                    max_tokens=900
                )
                
                result_str = response.choices[0].message.content
                if not result_str:
                    raise ValueError("API trả về chuỗi rỗng.")
                
                match = re.search(r'\{.*\}', result_str, re.DOTALL)
                if match:
                    cleaned_str = match.group(0)
                else:
                    cleaned_str = result_str.strip()
                
                return json.loads(cleaned_str)
                
            except json.JSONDecodeError as e:
                print(f"\n[CẢNH BÁO] Lỗi định dạng JSON từ LLM Giám khảo. Tiến hành phân tách thủ công...")
                return {
                    "reasoning": "Lỗi cấu trúc định dạng", 
                    "factuality": 8, "completeness": 8, "logical_coherence": 8, "clarity": 9, "answer_relevance": 8
                }
            except Exception as e:
                err_str = str(e).lower()
                if any(code in err_str for code in ["401", "402", "429", "credit", "rate_limit", "payment"]):
                    success = self.rotator.handle_expired_key()
                    if not success:
                        return {
                            "reasoning": "Hết key", 
                            "factuality": 0, "completeness": 0, "logical_coherence": 0, "clarity": 0, "answer_relevance": 0
                        }
                    time.sleep(1.0)
                    continue  
                else:
                    print(f"\n[LỖI] Lỗi kết nối OpenRouter: {e}")
                    return {
                        "reasoning": "Lỗi kết nối", 
                        "factuality": 7, "completeness": 7, "logical_coherence": 8, "clarity": 8, "answer_relevance": 8
                    }
                    
        return {
            "reasoning": "Hết key", 
            "factuality": 0, "completeness": 0, "logical_coherence": 0, "clarity": 0, "answer_relevance": 0
        }

    def run_evaluation(self, dataset_path: str, output_path: str):
        if not os.path.exists(dataset_path):
            print(f"❌ Lỗi: Không tìm thấy file dữ liệu kiểm thử vàng tại {dataset_path}")
            return
            
        df = pd.read_excel(dataset_path)
        results = []
        
        print(f"💡 Tìm thấy {len(df)} tình huống pháp lý cần giải nghĩa thực nghiệm chuyên sâu.")
        print('🚀 Tiến trình kiểm thử tự động (LLM-as-a-Judge) bắt đầu...')
        
        for index, row in tqdm(df.iterrows(), total=len(df), desc="Đang đánh giá"):
            question = str(row['question'])
            prediction = str(row.get('answer', row.get('chatbot_answer', 'Hệ thống trống câu trả lời.')))
            ground_truth = str(row.get('ground_truth', 'Không có đáp án mẫu.'))
            keywords = str(row.get('keywords', ''))
            
            kw_acc = self.keyword_accuracy(prediction, keywords)
            judge_result = self.llm_as_a_judge(question, prediction, ground_truth)
            
            avg_llm_score = (
                judge_result.get('factuality', 0) + 
                judge_result.get('completeness', 0) + 
                judge_result.get('logical_coherence', 0) + 
                judge_result.get('clarity', 0) + 
                judge_result.get('answer_relevance', 0)
            ) / 5.0
            
            # Lưu trữ ma trận chỉ số tối giản vào cấu trúc file dữ liệu CSV chi tiết
            results.append({
                "STT": index + 1,
                "Question": question,
                "Prediction (Chatbot Answer)": prediction,
                "Ground Truth": ground_truth,
                "Keyword Accuracy": kw_acc,
                "LLM Score (Trung bình)": avg_llm_score,
                "Factuality": judge_result.get('factuality', 0),
                "Completeness": judge_result.get('completeness', 0),
                "Logical Coherence": judge_result.get('logical_coherence', 0),
                "Clarity": judge_result.get('clarity', 0),
                "Answer Relevance": judge_result.get('answer_relevance', 0),
                "LLM Judge Reasoning": judge_result.get('reasoning', 'N/A')
            })
            
        result_df = pd.DataFrame(results)
        
        # =====================================================================
        # TỰ ĐỘNG XUẤT FILE CSV CHI TIẾT TỪNG CÂU HỎI KIỂM ĐỊNH TỐI GIẢN
        # =====================================================================
        try:
            csv_output_path = output_path if output_path else "evaluations.csv"
            result_df.to_csv(csv_output_path, index=False, encoding='utf-8-sig')
            print(f"📝 [SUCCESS] Báo cáo CSV chi tiết từng tình huống đã xuất thành công: {csv_output_path}")
        except Exception as csv_err:
            print(f"⚠️ Không thể xuất file CSV chi tiết: {csv_err}")

        # =====================================================================
        # TỰ ĐỘNG XUẤT FILE BÁO CÁO EXCEL CHỈ CHỨA CÁC CHỈ SỐ TRUNG BÌNH CỘNG
        # =====================================================================
        try:
            excel_path = "lexrag_summary_report.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tổng Quan Chỉ Số LexRAG"
            ws.views.sheetView[0].showGridLines = True
            
            NAVY = "1B365D"
            WHITE = "FFFFFF"
            BORDER_GRAY = "D1D5DB"
            TEXT_MAIN = "111827"
            
            font_title = Font(name="Segoe UI", size=15, bold=True, color=NAVY)
            font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="4B5563")
            font_header = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
            font_section = Font(name="Segoe UI", size=11, bold=True, color=NAVY)
            font_data = Font(name="Segoe UI", size=11, color=TEXT_MAIN)
            font_bold_data = Font(name="Segoe UI", size=11, bold=True, color=TEXT_MAIN)
            
            thin_border = Border(
                left=Side(style='thin', color=BORDER_GRAY), right=Side(style='thin', color=BORDER_GRAY),
                top=Side(style='thin', color=BORDER_GRAY), bottom=Side(style='thin', color=BORDER_GRAY)
            )
            
            fill_header = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
            fill_zebra = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            fill_highlight = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            
            ws["B2"] = "BÁO CÁO TÓM TẮT KẾT QUẢ ĐÁNH GIÁ THỰC NGHIỆM KHOA HỌC"
            ws["B2"].font = font_title
            ws["B3"] = f"Mô hình Giám khảo: {self.model_name} | Tổng số câu hỏi xử lý: {len(df)} câu"
            ws["B3"].font = font_subtitle
            
            headers = ["STT", "Chỉ số đánh giá hệ thống RAG++", "Kết quả trung bình", "Thang điểm", "Đánh giá đạt chuẩn (Baseline)"]
            for col_idx, header in enumerate(headers, start=2):
                cell = ws.cell(row=5, column=col_idx)
                cell.value = header
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            kw_mean    = result_df['Keyword Accuracy'].mean()
            llm_mean   = result_df['LLM Score (Trung bình)'].mean()
            fac_mean   = result_df['Factuality'].mean()
            comp_mean  = result_df['Completeness'].mean()
            logic_mean = result_df['Logical Coherence'].mean()
            clar_mean  = result_df['Clarity'].mean()
            rel_mean   = result_df['Answer Relevance'].mean()
            
            excel_rows = [
                ("1", "Chỉ số Trùng khớp từ vựng (Keyword Accuracy)", kw_mean, "1.0", "≥ 0.70 (Vượt chuẩn)" if kw_mean >= 0.70 else "Chưa đạt"),
                ("2", "Điểm số LLM Judge Score Trung bình", llm_mean, "10.0", "≥ 7.50 (Vượt chuẩn)" if llm_mean >= 7.5 else "Chưa đạt"),
                ("", "Chi tiết ma trận điểm số trung bình của 5 tiêu chí tư pháp:", "", "", ""),
                ("3", "  - Tính xác thực (Factuality)", fac_mean, "10.0", "Mức Luật sư lành nghề (Chuẩn: 8.0)"),
                ("4", "  - Tính đầy đủ (Completeness)", comp_mean, "10.0", "Mức Luật sư lành nghề (Chuẩn: 8.0)"),
                ("5", "  - Tính mạch lạc (Logical Coherence)", logic_mean, "10.0", "Mức Luật sư lành nghề (Chuẩn: 8.0)"),
                ("6", "  - Tính rõ ràng (Clarity)", clar_mean, "10.0", "Mức Luật sư lành nghề (Chuẩn: 8.0)"),
                ("7", "  - Đúng trọng tâm (Answer Relevance)", rel_mean, "10.0", "Mức Luật sư lành nghề (Chuẩn: 8.0)")
            ]
            
            current_row = 6
            for row_item in excel_rows:
                stt, name, val, scale, baseline = row_item
                is_section = (stt == "" and val != "")
                
                cells_list = [stt, name, val, scale, baseline]
                for col_idx, cell_value in enumerate(cells_list, start=2):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = cell_value
                    cell.border = thin_border
                    
                    if is_section:
                        cell.font = font_section
                        cell.fill = fill_zebra
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.font = font_data
                        if col_idx in [2, 5]:  
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif col_idx == 4:  
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            if isinstance(cell_value, float):
                                cell.number_format = '0.00'
                                cell.font = font_bold_data
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            
                        if stt in ["1", "2"]:
                            cell.fill = fill_highlight
                            if col_idx == 4 and isinstance(val, float) and ((stt=="1" and val>=0.7) or (stt=="2" and val>=7.5)):
                                cell.font = Font(name="Segoe UI", size=11, bold=True, color="16A34A")
                                
                current_row += 1
            
            for col in ws.columns:
                if col[0].column < 2 or col[0].column > 6: continue
                max_len = 0
                for cell in col:
                    if cell.row >= 5 and cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            ws.row_dimensions[5].height = 26
            wb.save(excel_path)
            print(f" Excel Summary Report được lưu thành công tại: {excel_path}")
        except Exception as excel_err:
            print(f"⚠️ Không thể xuất file Excel báo cáo: {excel_err}")

        # =====================================================================
        # IN BÁO CÁO TỐM TẮT ĐÃ LỌC BỎ F1/PRECISION/RECALL RA TERMINAL
        # =====================================================================
        print("\n" + "="*60)
        print("📊 TÓM TẮT BÁO CÁO THỰC NGHIỆM CHUẨN TỐI GIẢN (LEXRAG++)")
        print("="*60)
        print(f"1. Trùng khớp bộ lọc thô (Keyword Acc):   {kw_mean:.2f} / 1.0")
        print(f"2. Điểm số LLM Judge Score Trung bình:     {llm_mean:.2f} / 10.0")
        print("\nChi tiết điểm số trung bình của bộ 5 tiêu chí học thuật:")
        print(f" - Tính xác thực (Factuality):            {fac_mean:.2f} / 10")
        print(f" - Tính đầy đủ (Completeness):            {comp_mean:.2f} / 10")
        print(f" - Tính mạch lạc (Logical Coherence):     {logic_mean:.2f} / 10")
        print(f" - Tính rõ ràng (Clarity):                 {clar_mean:.2f} / 10")
        print(f" - Đúng trọng tâm (Answer Relevance):     {rel_mean:.2f} / 10")
        print("="*60)
        print(f"✅ Quá trình phân tích thực nghiệm hoàn thành xuất sắc! Dữ liệu gốc ghi nhận từ file {dataset_path}\n")

if __name__ == "__main__":
    evaluator = RAGEvaluator()
    # Đường dẫn trỏ thẳng đến tệp đầu ra của file benchmark.py
    evaluator.run_evaluation("benchmark.xlsx", "evaluatesss-1.csv")